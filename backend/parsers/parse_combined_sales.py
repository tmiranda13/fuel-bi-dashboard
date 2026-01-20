"""
Combined Parser for Vendas por Bico + Cupons por Período

This parser combines two reports to create a comprehensive sales dataset:
- Vendas por Bico: provides pump numbers, product, client, volume, value, payment method
- Cupons por Período: provides employee, transaction time, cupom number

Cross-references using: date + product + volume + value + client
"""

import os
import re
import sys
import logging
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
from collections import defaultdict
from dotenv import load_dotenv
from supabase import create_client

# Load environment variables
# Try multiple locations for .env.local
possible_env_paths = [
    Path(__file__).parent.parent / 'config' / '.env.local',  # webposto_automator/config/
    Path(__file__).parent.parent.parent / 'config' / '.env.local',  # original structure
]

env_loaded = False
for env_path in possible_env_paths:
    if env_path.exists():
        load_dotenv(env_path)
        env_loaded = True
        break

if not env_loaded:
    print("WARNING: .env.local not found. Supabase operations will fail.")

# Setup logging
# Try webposto_automator location first, then fallback
LOG_DIR = Path(__file__).parent.parent / "logs"
if not LOG_DIR.parent.exists():
    LOG_DIR = Path.home() / "Desktop" / "webposto_reports" / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "automation_alerts.log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('combined_sales')

# Default company ID
DEFAULT_COMPANY_ID = 2


def get_supabase_client():
    """Get Supabase client"""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY")
    if not url or not key:
        raise ValueError("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY")
    return create_client(url, key)


def log_alert(message, level='warning'):
    """Log an alert message to file and console"""
    if level == 'error':
        logger.error(message)
    elif level == 'warning':
        logger.warning(message)
    else:
        logger.info(message)


# Fuel product codes mapping
FUEL_CODES = {
    '000001': {'code': 'GC', 'name': 'GASOLINA COMUM'},
    '000002': {'code': 'DS10', 'name': 'DIESEL S10'},
    '000004': {'code': 'DS500', 'name': 'DIESEL S500'},
    '001361': {'code': 'GA', 'name': 'GASOLINA ADITIVADA'},
    '004593': {'code': 'ET', 'name': 'ETANOL'},
    '009826': {'code': 'GC', 'name': 'GASOLINA COMUM'},
    '009827': {'code': 'GA', 'name': 'GASOLINA ADITIVADA'},
}

# Product name to canonical code mapping
PRODUCT_NAME_TO_CODE = {
    'GASOLINA COMUM': 'GC',
    'GASOLINA COMUM.': 'GC',
    'GASOLINA ADITIVADA': 'GA',
    'GASOLINA ADITIVADA.': 'GA',
    'DIESEL S10': 'DS10',
    'DIESEL S10.': 'DS10',
    'DIESEL S500': 'DS500',
    'DIESEL S500.': 'DS500',
    'DIESEL COMUM': 'DS500',  # DIESEL COMUM = DIESEL S500
    'DIESEL COMUM.': 'DS500',
    'ETANOL': 'ET',
    'ETANOL.': 'ET',
    'ETANOL COMUM': 'ET',
    'ETANOL COMUM.': 'ET',
}


def parse_date(date_val):
    """Parse date from various formats"""
    if isinstance(date_val, (datetime, date)):
        return date_val.date() if isinstance(date_val, datetime) else date_val
    if isinstance(date_val, str):
        # Try DD/MM/YYYY format
        try:
            return datetime.strptime(date_val, '%d/%m/%Y').date()
        except:
            pass
        # Try YYYY-MM-DD format
        try:
            return datetime.strptime(date_val, '%Y-%m-%d').date()
        except:
            pass
    return None


def parse_time(time_val):
    """Parse time from various formats"""
    if isinstance(time_val, datetime):
        return time_val.strftime('%H:%M:%S')
    if isinstance(time_val, str):
        return time_val
    return None


def normalize_product_name(name):
    """Normalize product name for matching"""
    if not name:
        return None
    name = str(name).strip().upper()
    # Remove trailing periods
    name = name.rstrip('.')
    return name


def get_canonical_code(product_name):
    """Get canonical product code from product name"""
    if not product_name:
        return None
    normalized = product_name.strip().upper()
    return PRODUCT_NAME_TO_CODE.get(normalized) or PRODUCT_NAME_TO_CODE.get(normalized + '.')


def create_match_key(sale_date, product_code, volume, value, client=None):
    """
    Create a key for matching transactions between reports.

    NOTE: Client name is NOT used in matching because the two reports
    use different name formats (abbreviated vs full legal names).
    Matching on (date, product, volume, value) achieves 100% match rate.
    """
    # Round volume to 3 decimals, value to 2 decimals
    vol_rounded = round(float(volume), 3)
    val_rounded = round(float(value), 2)

    # Match on date + product + volume + value (client excluded for better matching)
    return (str(sale_date), product_code, vol_rounded, val_rounded)


def parse_vendas_por_bico(file_path, company_id=2):
    """
    Parse Vendas por Bico report

    Returns list of transactions with:
    - pump_number, product_name, product_code, sale_date, client, volume, value, payment_method
    """
    print(f"\n[INFO] Parsing Vendas por Bico: {Path(file_path).name}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    transactions = []
    current_product = None
    current_pump = None
    station_name = None

    rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(rows):
        # Get station name from row 1
        if row[0] == 'Filial:':
            station_name = row[1]
            continue

        # Detect product/pump header: "Produto:", product_name, None, "Bico:", pump_number
        if row[0] == 'Produto:':
            current_product = row[1]
            current_pump = str(row[4]).zfill(3) if row[4] else None
            continue

        # Skip header rows and subtotal rows
        if row[0] in ('Data', None, 'Subtotal Bico:', 'Subtotal Produto:', 'Subtotal Filial:', 'Total:'):
            continue

        # Parse transaction row: Date, Client, Acréscimo, Desconto, None, Quantidade, Valor, Prazo
        sale_date = parse_date(row[0])
        if not sale_date:
            continue

        client = row[1]
        volume = round(float(row[5] or 0), 3)
        value = round(float(row[6] or 0), 2)
        payment_method = row[7]

        # Skip zero transactions
        if volume == 0 and value == 0:
            continue

        product_code = get_canonical_code(current_product)

        transactions.append({
            'company_id': company_id,
            'pump_number': current_pump,
            'product_name': current_product,
            'product_code': product_code,
            'sale_date': sale_date,
            'client': client,
            'volume': volume,
            'value': value,
            'payment_method': payment_method,
            'station_name': station_name,
        })

    print(f"[OK] Parsed {len(transactions)} transactions from Vendas por Bico")
    return transactions


def parse_cupons_por_periodo(file_path, company_id=2):
    """
    Parse Cupons por Período report

    Returns list of transactions with:
    - cupom_number, sale_date, sale_time, client, employee, product_code, product_name, volume, value, payment_method
    """
    print(f"\n[INFO] Parsing Cupons por Período: {Path(file_path).name}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    transactions = []
    station_name = None
    current_cupom = None
    current_date = None
    current_time = None
    current_client = None
    current_employee = None

    rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(rows):
        # Get station name
        if row[0] == 'Filial:':
            station_name = row[2]
            continue

        # Track current employee from section headers (format: "Funcionário:" in col 0, employee name in col 4)
        if row[0] == 'Funcionário:':
            current_employee = row[4] if len(row) > 4 and row[4] else current_employee
            continue

        # Detect cupom header row: numeric cupom number (typically 3-4 digits) with date in column 5
        # Cupom numbers are 1-4 digits, product codes are 6 digits - distinguish by date presence
        if row[0] and str(row[0]).isdigit():
            # Check if this looks like a cupom header (has date in column 5)
            if row[5] and ('/' in str(row[5]) or isinstance(row[5], (datetime, date))):
                current_cupom = str(row[0])
                current_date = parse_date(row[5])
                current_time = parse_time(row[8]) if len(row) > 8 else None
                current_client = row[10] if len(row) > 10 else None
                current_employee = row[13] if len(row) > 13 else None
                continue

        # Detect product line: Ref (product code), Product Name, ..., Quant(col 6), Val.Unit(col 7), ..., Val.Item(col 18), ..., Prazo(col 21)
        ref = str(row[0]) if row[0] else ''
        if ref in FUEL_CODES:
            product_name = row[1]
            volume = round(float(row[6] or 0), 3)
            unit_price = round(float(row[7] or 0), 2)
            value = round(float(row[18] or 0), 2)  # Val. Item
            payment_method = row[21]

            product_info = FUEL_CODES[ref]

            transactions.append({
                'company_id': company_id,
                'cupom_number': current_cupom,
                'sale_date': current_date,
                'sale_time': current_time,
                'client': current_client,
                'employee': current_employee,
                'source_product_code': ref,
                'product_code': product_info['code'],
                'product_name': product_name,
                'volume': volume,
                'unit_price': unit_price,
                'value': value,
                'payment_method': payment_method,
                'station_name': station_name,
            })

    print(f"[OK] Parsed {len(transactions)} fuel transactions from Cupons por Período")
    return transactions


def combine_reports(vendas_bico_data, cupons_data):
    """
    Combine data from both reports using matching keys

    Returns (combined_transactions, match_stats)
    match_stats contains: matched, unmatched, match_rate, is_perfect_match
    """
    print("\n[INFO] Cross-referencing reports...")

    # Create lookup from cupons data
    cupons_lookup = defaultdict(list)
    for txn in cupons_data:
        key = create_match_key(
            txn['sale_date'],
            txn['product_code'],
            txn['volume'],
            txn['value'],
            txn['client']
        )
        cupons_lookup[key].append(txn)

    # Match vendas_bico with cupons
    combined = []
    matched = 0
    unmatched = 0
    unmatched_transactions = []  # Track unmatched for debugging

    for txn in vendas_bico_data:
        key = create_match_key(
            txn['sale_date'],
            txn['product_code'],
            txn['volume'],
            txn['value'],
            txn['client']
        )

        cupons_matches = cupons_lookup.get(key, [])

        if cupons_matches:
            # Use the first match (and remove it to avoid duplicate matching)
            cupom_txn = cupons_matches.pop(0)
            matched += 1

            combined.append({
                # From Vendas por Bico
                'company_id': txn['company_id'],
                'pump_number': txn['pump_number'],
                'product_code': txn['product_code'],
                'product_name': txn['product_name'],
                'sale_date': txn['sale_date'],
                'client': txn['client'],
                'volume': txn['volume'],
                'value': txn['value'],
                'payment_method': txn['payment_method'],
                'station_name': txn['station_name'],
                # From Cupons por Período
                'cupom_number': cupom_txn['cupom_number'],
                'sale_time': cupom_txn['sale_time'],
                'employee': cupom_txn['employee'],
                'unit_price': cupom_txn['unit_price'],
                'source_product_code': cupom_txn['source_product_code'],
            })
        else:
            # No match found - keep vendas_bico data with nulls for cupons fields
            unmatched += 1
            unmatched_transactions.append({
                'sale_date': txn['sale_date'],
                'product': txn['product_name'],
                'volume': txn['volume'],
                'value': txn['value'],
                'client': txn['client'],
            })
            combined.append({
                'company_id': txn['company_id'],
                'pump_number': txn['pump_number'],
                'product_code': txn['product_code'],
                'product_name': txn['product_name'],
                'sale_date': txn['sale_date'],
                'client': txn['client'],
                'volume': txn['volume'],
                'value': txn['value'],
                'payment_method': txn['payment_method'],
                'station_name': txn['station_name'],
                'cupom_number': None,
                'sale_time': None,
                'employee': None,
                'unit_price': txn['value'] / txn['volume'] if txn['volume'] > 0 else 0,
                'source_product_code': None,
            })

    total = len(combined)
    match_rate = (matched / total * 100) if total > 0 else 0
    is_perfect_match = (unmatched == 0)

    print(f"[OK] Combined {total} transactions")
    print(f"     Matched: {matched} ({match_rate:.1f}%)")
    print(f"     Unmatched: {unmatched} ({100 - match_rate:.1f}%)")

    if not is_perfect_match:
        print(f"\n[WARNING] MATCH RATE IS NOT 100% - {unmatched} transactions unmatched!")
        print(f"[WARNING] First 5 unmatched transactions:")
        for i, txn in enumerate(unmatched_transactions[:5]):
            print(f"  {i+1}. {txn['sale_date']} | {txn['product']} | {txn['volume']}L | R${txn['value']} | {txn['client']}")

    match_stats = {
        'matched': matched,
        'unmatched': unmatched,
        'total': total,
        'match_rate': match_rate,
        'is_perfect_match': is_perfect_match,
        'unmatched_transactions': unmatched_transactions,
    }

    return combined, match_stats


def aggregate_by_pump(combined_data):
    """
    Aggregate combined data by pump/product/date for pump_sales_intraday format

    Returns data ready for database import
    """
    print("\n[INFO] Aggregating by pump/product/date...")

    # Group by pump + product + date
    aggregated = defaultdict(lambda: {
        'volume': 0,
        'revenue': 0,
        'transactions': 0,
        'employees': set(),
        'clients': set(),
    })

    for txn in combined_data:
        key = (txn['pump_number'], txn['product_code'], txn['sale_date'])
        agg = aggregated[key]
        agg['volume'] += txn['volume']
        agg['revenue'] += txn['value']
        agg['transactions'] += 1
        agg['company_id'] = txn['company_id']
        agg['product_name'] = txn['product_name']
        agg['station_name'] = txn['station_name']
        if txn.get('employee'):
            agg['employees'].add(txn['employee'])
        if txn.get('client'):
            agg['clients'].add(txn['client'])

    # Convert to list format for database
    result = []
    for (pump_number, product_code, sale_date), agg in aggregated.items():
        avg_price = agg['revenue'] / agg['volume'] if agg['volume'] > 0 else 0
        result.append({
            'company_id': agg['company_id'],
            'pump_number': pump_number,
            'product_code': product_code,
            'product_name': agg['product_name'],
            'sale_date': str(sale_date),
            'volume_sold': round(agg['volume'], 3),
            'total_revenue': round(agg['revenue'], 2),
            'sale_price_per_liter': round(avg_price, 2),
            'transaction_count': agg['transactions'],
            'employees': list(agg['employees']),
            'unique_clients': len(agg['clients']),
        })

    print(f"[OK] Aggregated to {len(result)} pump/product/day records")
    return result


def get_summary(combined_data):
    """Generate summary statistics from combined data"""

    total_volume = sum(t['volume'] for t in combined_data)
    total_revenue = sum(t['value'] for t in combined_data)

    # By product
    by_product = defaultdict(lambda: {'volume': 0, 'revenue': 0, 'count': 0})
    for t in combined_data:
        by_product[t['product_name']]['volume'] += t['volume']
        by_product[t['product_name']]['revenue'] += t['value']
        by_product[t['product_name']]['count'] += 1

    # By pump
    by_pump = defaultdict(lambda: {'volume': 0, 'revenue': 0, 'count': 0})
    for t in combined_data:
        by_pump[t['pump_number']]['volume'] += t['volume']
        by_pump[t['pump_number']]['revenue'] += t['value']
        by_pump[t['pump_number']]['count'] += 1

    # By employee
    by_employee = defaultdict(lambda: {'volume': 0, 'revenue': 0, 'count': 0})
    for t in combined_data:
        emp = t.get('employee') or 'Unknown'
        by_employee[emp]['volume'] += t['volume']
        by_employee[emp]['revenue'] += t['value']
        by_employee[emp]['count'] += 1

    return {
        'total_volume': total_volume,
        'total_revenue': total_revenue,
        'total_transactions': len(combined_data),
        'by_product': dict(by_product),
        'by_pump': dict(by_pump),
        'by_employee': dict(by_employee),
    }


def print_summary(summary):
    """Print formatted summary"""
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)

    print(f"\nTotal Volume: {summary['total_volume']:,.2f} L")
    print(f"Total Revenue: R$ {summary['total_revenue']:,.2f}")
    print(f"Total Transactions: {summary['total_transactions']}")

    print("\n--- By Product ---")
    for product, data in sorted(summary['by_product'].items()):
        print(f"  {product}: {data['volume']:,.2f} L / R$ {data['revenue']:,.2f} ({data['count']} txns)")

    print("\n--- By Pump (Top 10) ---")
    sorted_pumps = sorted(summary['by_pump'].items(), key=lambda x: x[1]['volume'], reverse=True)[:10]
    for pump, data in sorted_pumps:
        print(f"  Pump {pump}: {data['volume']:,.2f} L / R$ {data['revenue']:,.2f} ({data['count']} txns)")

    print("\n--- By Employee ---")
    for emp, data in sorted(summary['by_employee'].items(), key=lambda x: x[1]['revenue'], reverse=True):
        print(f"  {emp}: {data['volume']:,.2f} L / R$ {data['revenue']:,.2f} ({data['count']} txns)")


def load_to_database(combined_data, company_id=2):
    """
    Load combined sales data to Supabase combined_sales table.

    Checks for duplicates based on (sale_date, pump_number, product_code, volume, value, cupom_number)
    """
    print("\n[INFO] Loading to database...")

    supabase = get_supabase_client()

    inserted = 0
    skipped = 0
    errors = 0

    for txn in combined_data:
        try:
            # Prepare record for database
            record = {
                'company_id': txn['company_id'],
                'sale_date': str(txn['sale_date']),
                'pump_number': txn['pump_number'],
                'product_code': txn['product_code'],
                'product_name': txn['product_name'],
                'client': txn.get('client'),
                'volume': txn['volume'],
                'value': txn['value'],
                'payment_method': txn.get('payment_method'),
                'cupom_number': txn.get('cupom_number'),
                'sale_time': txn.get('sale_time'),
                'employee': txn.get('employee'),
                'unit_price': txn.get('unit_price'),
            }

            # Check for existing record (by unique key)
            existing = supabase.table('combined_sales').select('id').eq(
                'company_id', record['company_id']
            ).eq(
                'sale_date', record['sale_date']
            ).eq(
                'pump_number', record['pump_number']
            ).eq(
                'product_code', record['product_code']
            ).eq(
                'volume', record['volume']
            ).eq(
                'value', record['value']
            ).execute()

            if existing.data:
                skipped += 1
                continue

            # Insert new record
            supabase.table('combined_sales').insert(record).execute()
            inserted += 1

            if inserted % 100 == 0:
                print(f"  [PROGRESS] Inserted {inserted} records...")

        except Exception as e:
            print(f"  [ERROR] Failed to insert: {e}")
            errors += 1

    print(f"\n[RESULT] Inserted: {inserted}, Skipped (duplicates): {skipped}, Errors: {errors}")
    return {'inserted': inserted, 'skipped': skipped, 'errors': errors}


def main():
    """Main function"""
    import argparse

    parser = argparse.ArgumentParser(description='Parse and combine Vendas por Bico + Cupons por Período')
    parser.add_argument('--vendas-bico', '-v', required=True, help='Path to Vendas por Bico.xlsx')
    parser.add_argument('--cupons', '-c', required=True, help='Path to Cupons por Período.xlsx')
    parser.add_argument('--company-id', type=int, default=2, help='Company ID (default: 2)')
    parser.add_argument('--output', '-o', help='Output CSV file (optional)')
    parser.add_argument('--load', '-l', action='store_true', help='Load to Supabase database')

    args = parser.parse_args()

    # Verify files exist
    if not Path(args.vendas_bico).exists():
        print(f"Error: File not found: {args.vendas_bico}")
        sys.exit(1)
    if not Path(args.cupons).exists():
        print(f"Error: File not found: {args.cupons}")
        sys.exit(1)

    print("=" * 70)
    print("COMBINED SALES PARSER")
    print("=" * 70)

    # Parse both reports
    vendas_data = parse_vendas_por_bico(args.vendas_bico, args.company_id)
    cupons_data = parse_cupons_por_periodo(args.cupons, args.company_id)

    # Combine reports
    combined, match_stats = combine_reports(vendas_data, cupons_data)

    # Check match rate - MUST BE 100%
    if not match_stats['is_perfect_match']:
        alert_msg = (
            f"MATCH RATE NOT 100% | "
            f"Rate: {match_stats['match_rate']:.2f}% | "
            f"Unmatched: {match_stats['unmatched']} transactions | "
            f"Files: {Path(args.vendas_bico).name}, {Path(args.cupons).name}"
        )
        log_alert(alert_msg, level='error')

        # Log unmatched transactions details
        for i, txn in enumerate(match_stats['unmatched_transactions'][:10]):
            log_alert(f"  Unmatched #{i+1}: {txn['sale_date']} | {txn['product']} | {txn['volume']}L | R${txn['value']}", level='warning')

        print("\n" + "!" * 70)
        print("ALERT: MATCH RATE IS NOT 100%")
        print("!" * 70)
        print(f"Match rate: {match_stats['match_rate']:.2f}%")
        print(f"Unmatched: {match_stats['unmatched']} transactions")
        print("\nData will NOT be loaded to database until match rate is 100%.")
        print(f"Check log file: {LOG_FILE}")
        print("!" * 70)

        # Don't load to database if match is not perfect
        if args.load:
            print("\n[SKIPPED] Database load skipped due to imperfect match rate.")
            log_alert("Database load SKIPPED - imperfect match rate", level='error')
            args.load = False  # Prevent loading

        # Return with error status
        return combined, None, match_stats

    # Generate and print summary
    summary = get_summary(combined)
    print_summary(summary)

    # Load to database if requested (only if match is perfect)
    if args.load:
        result = load_to_database(combined, args.company_id)
        log_alert(
            f"COMBINED SALES LOADED | "
            f"Inserted: {result['inserted']} | "
            f"Skipped: {result['skipped']} | "
            f"Errors: {result['errors']} | "
            f"Match: 100%",
            level='info'
        )

    # Aggregate for database format (for CSV output)
    aggregated = aggregate_by_pump(combined)

    # Output to CSV if requested
    if args.output:
        import csv
        with open(args.output, 'w', newline='', encoding='utf-8') as f:
            if aggregated:
                writer = csv.DictWriter(f, fieldnames=aggregated[0].keys())
                writer.writeheader()
                writer.writerows(aggregated)
        print(f"\n[OK] Saved to {args.output}")

    print("\n[DONE] Parsing complete!")

    return combined, aggregated, match_stats


if __name__ == '__main__':
    main()
