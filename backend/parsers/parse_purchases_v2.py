"""
Parser for Entrada de Mercadorias V2 (with supplier info in header rows)

New format structure:
- Row with "Cliente/Fornecedor: XXXXX - SUPPLIER NAME" = supplier header
- Next row = column headers (Ref., Data Emissão, etc.)
- Following rows = product data until empty row or next supplier header
- Repeats for each supplier
- Ends with "Total Geral:" row

Extracts supplier_code and supplier_name from header rows and associates with each purchase.
"""

import os
import re
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
from supabase import create_client

# =============================================================================
# FUEL PRODUCTS FILTER
# Only these source codes will be imported (non-fuel products are skipped)
# =============================================================================
FUEL_SOURCE_CODES = {
    '000001': 'GC',      # GASOLINA COMUM
    '000002': 'DS10',    # DIESEL S10
    '000004': 'DS500',   # DIESEL S500
    '001361': 'GA',      # GASOLINA ADITIVADA
    '004593': 'ET',      # ETANOL
    '009826': 'GC',      # GASOLINA COMUM.
    '009827': 'GA',      # GASOLINA ADITIVADA.
}

# Load environment variables
# Path: parsers -> backend -> 1strev -> config
env_path = Path(__file__).parent.parent.parent / 'config' / '.env.local'
if env_path.exists():
    load_dotenv(env_path)
else:
    # Try alternate path (backend -> config)
    env_path = Path(__file__).parent.parent / 'config' / '.env.local'
    if env_path.exists():
        load_dotenv(env_path)
    else:
        print(f"Warning: Could not find .env.local file")

def get_supabase_client():
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY")
    if not url or not key:
        raise ValueError("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY environment variables")
    return create_client(url, key)

def get_product_code_mapping(supabase, company_id):
    """Get product code mapping from database"""
    result = supabase.table('product_code_mapping').select('*').eq('company_id', company_id).execute()
    mapping = {}
    for row in result.data:
        mapping[row['source_product_code']] = {
            'canonical_code': row['canonical_product_code'],
            'canonical_name': row.get('source_product_name', row['canonical_product_code'])  # Use source_product_name as display name
        }
    return mapping

def parse_supplier_header(cell_value):
    """
    Parse supplier header like "Cliente/Fornecedor: 001436 - RAIZEN S.A. 5"
    Returns (supplier_code, supplier_name) or (None, None)
    """
    if not cell_value or 'Cliente/Fornecedor:' not in str(cell_value):
        return None, None
    
    # Extract the part after "Cliente/Fornecedor:"
    text = str(cell_value).replace('Cliente/Fornecedor:', '').strip()
    
    # Split by " - " to separate code and name
    parts = text.split(' - ', 1)
    if len(parts) == 2:
        supplier_code = parts[0].strip()
        supplier_name = parts[1].strip()
        return supplier_code, supplier_name
    
    return None, text  # Return full text as name if no code found

def parse_purchases_v2(file_path, company_id):
    """
    Parse the V2 format of Entrada de Mercadorias Excel file.
    Returns list of purchase records with supplier info.
    """
    print(f"Parsing file: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    records = []
    current_supplier_code = None
    current_supplier_name = None
    skip_next_row = False  # Flag to skip column header rows
    
    for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        # Get first cell value
        first_cell = row[0] if row else None
        
        # Skip empty rows
        if not first_cell or str(first_cell).strip() == '':
            skip_next_row = False
            continue
        
        first_cell_str = str(first_cell).strip()
        
        # Check for supplier header
        if 'Cliente/Fornecedor:' in first_cell_str:
            current_supplier_code, current_supplier_name = parse_supplier_header(first_cell_str)
            print(f"  Found supplier: {current_supplier_code} - {current_supplier_name}")
            skip_next_row = True  # Next row will be column headers
            continue
        
        # Skip column header rows (Ref., Data Emissão, etc.)
        if first_cell_str == 'Ref.' or skip_next_row:
            skip_next_row = False
            continue
        
        # Skip total row
        if 'Total' in first_cell_str:
            continue
        
        # This should be a data row - parse it
        try:
            # Column mapping for V2 format:
            # Col 0: Ref (source_product_code)
            # Col 3: Data Emissão (invoice_date)
            # Col 5: Descrição (product_name)
            # Col 7: Entrada (receipt_date)
            # Col 9: Qtde (quantity)
            # Col 10: P. Custo R$ (cost_price)
            # Col 12: Valor Unit. R$ (unit_value)
            # Col 13: P. Venda R$ (selling_price)
            # Col 15: Mkp % (markup_percentage)
            # Col 17: Subtotal R$ (subtotal)
            # Col 18: NFe (invoice_number)
            # Col 19: Estoque (warehouse)
            
            source_product_code = str(row[0]).strip() if row[0] else None
            
            # Skip if no product code
            if not source_product_code:
                continue
            
            # Parse dates
            invoice_date = row[3]
            if isinstance(invoice_date, datetime):
                invoice_date = invoice_date.date()
            elif isinstance(invoice_date, str):
                invoice_date = datetime.strptime(invoice_date.split()[0], '%Y-%m-%d').date()
            
            receipt_date = row[7]
            if isinstance(receipt_date, datetime):
                receipt_date = receipt_date.date()
            elif isinstance(receipt_date, str):
                receipt_date = datetime.strptime(receipt_date.split()[0], '%Y-%m-%d').date()
            
            record = {
                'company_id': company_id,
                'source_product_code': source_product_code,
                'source_product_name': str(row[5]).strip() if row[5] else None,
                'invoice_date': str(invoice_date),
                'receipt_date': str(receipt_date),
                'quantity': float(row[9]) if row[9] else 0,
                'cost_price': float(row[10]) if row[10] else 0,
                'unit_value': float(row[12]) if row[12] else 0,
                'selling_price': float(row[13]) if row[13] else 0,
                'markup_percentage': float(row[15]) if row[15] else 0,
                'subtotal': float(row[17]) if row[17] else 0,
                'invoice_number': str(row[18]).strip() if row[18] else None,
                'warehouse': str(row[19]).strip() if row[19] else None,
                'supplier_code': current_supplier_code,
                'supplier_name': current_supplier_name,
                'source_file_name': Path(file_path).name
            }
            
            records.append(record)
            
        except Exception as e:
            print(f"  Warning: Error parsing row {row_num}: {e}")
            continue
    
    wb.close()
    print(f"  Parsed {len(records)} purchase records")
    return records

def load_to_database(records, supabase, company_id):
    """Load parsed records to Supabase database (FUEL PRODUCTS ONLY)"""

    # Get product code mapping from database (optional override)
    code_mapping = get_product_code_mapping(supabase, company_id)

    inserted = 0
    skipped = 0
    filtered = 0
    errors = 0

    for record in records:
        try:
            # FUEL FILTER: Only process fuel products
            source_code = record['source_product_code']
            if source_code not in FUEL_SOURCE_CODES:
                filtered += 1
                continue

            # Map source code to canonical code using hardcoded FUEL_SOURCE_CODES
            canonical_code = FUEL_SOURCE_CODES[source_code]
            record['canonical_product_code'] = canonical_code
            record['product_name'] = record['source_product_name']

            # Override with database mapping if available
            if source_code in code_mapping:
                record['canonical_product_code'] = code_mapping[source_code]['canonical_code']
                record['product_name'] = code_mapping[source_code]['canonical_name']
            
            # Check for existing record (by invoice_number, receipt_date, source_product_code)
            existing = supabase.table('purchases').select('id').eq(
                'company_id', company_id
            ).eq(
                'invoice_number', record['invoice_number']
            ).eq(
                'receipt_date', record['receipt_date']
            ).eq(
                'source_product_code', record['source_product_code']
            ).execute()
            
            if existing.data:
                # Update existing record with supplier info
                supabase.table('purchases').update({
                    'supplier_code': record['supplier_code'],
                    'supplier_name': record['supplier_name']
                }).eq('id', existing.data[0]['id']).execute()
                skipped += 1
                continue
            
            # Insert new record
            supabase.table('purchases').insert(record).execute()
            inserted += 1
            
        except Exception as e:
            print(f"  Error inserting record: {e}")
            errors += 1
    
    print(f"  Results: {inserted} inserted, {skipped} updated, {filtered} non-fuel filtered, {errors} errors")
    return inserted, skipped, errors

def main():
    """Main function to run the parser"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Parse Entrada de Mercadorias V2 Excel files')
    parser.add_argument('file_path', nargs='?', help='Path to Excel file (optional, uses default if not provided)')
    parser.add_argument('--company-id', type=int, default=1, help='Company ID (default: 1)')
    
    args = parser.parse_args()
    
    # Default file path if not provided
    if args.file_path:
        file_path = args.file_path
    else:
        # Look for Entrada de mercadorias_*.xlsx files in docs folder
        # Path: parsers -> backend -> 1strev -> docs
        docs_path = Path(__file__).parent.parent.parent / 'docs'
        print(f"Looking in: {docs_path}")
        entrada_files = list(docs_path.glob('Entrada de mercadorias_*.xlsx'))
        if entrada_files:
            # Sort to get consistent order and process all files
            entrada_files.sort()
            print(f"Found {len(entrada_files)} file(s) to process")
            
            supabase = get_supabase_client()
            for f in entrada_files:
                print(f"\n{'='*50}")
                print(f"Processing: {f.name}")
                records = parse_purchases_v2(str(f), args.company_id)
                if records:
                    load_to_database(records, supabase, args.company_id)
            
            print("\nAll files processed!")
            sys.exit(0)
        else:
            print("No 'Entrada de mercadorias_*.xlsx' files found in docs folder.")
            print("Usage: python parse_purchases_v2.py <file_path> [--company-id N]")
            sys.exit(1)
    
    # Verify file exists
    if not Path(file_path).exists():
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    # Parse the file
    records = parse_purchases_v2(file_path, args.company_id)
    
    if not records:
        print("No records parsed from file")
        sys.exit(1)
    
    # Load to database
    print("\nLoading to database...")
    supabase = get_supabase_client()
    load_to_database(records, supabase, args.company_id)
    
    print("\nDone!")

if __name__ == '__main__':
    main()
