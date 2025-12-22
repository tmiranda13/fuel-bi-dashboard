"""
Parser for Inventory Variance (Estoque de Combustível section from RESUMO DO DIA)

Extracts the "Estoque de Combustível" section which shows:
- Inicial: Initial stock
- Final: Calculated final stock
- Medição: Physical measurement
- Diferença: Variance (measurement - final)

This data helps identify losses, gains, or measurement errors.
"""

import os
import re
import sys
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
from dotenv import load_dotenv
from supabase import create_client

# Load environment variables
# Path: parsers -> backend -> 1strev -> config
env_path = Path(__file__).parent.parent.parent / 'config' / '.env.local'
if env_path.exists():
    load_dotenv(env_path)
else:
    env_path = Path(__file__).parent.parent / 'config' / '.env.local'
    if env_path.exists():
        load_dotenv(env_path)
    else:
        print(f"Warning: Could not find .env.local file")

# Product name to code mapping
PRODUCT_CODE_MAP = {
    'DIESEL COMUM': 'DS500',
    'DIESEL S500': 'DS500',
    'DIESEL S10': 'DS10',
    'ETANOL': 'ET',
    'GASOLINA COMUM': 'GC',
    'GASOLINA ADITIVADA': 'GA',
}

def get_supabase_client():
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY")
    if not url or not key:
        raise ValueError("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY environment variables")
    return create_client(url, key)

def extract_date_from_filename(filename):
    """
    Extract date from filename like 'RESUMO DO DIA_9_18_2025.xlsx'
    Returns date object or None
    """
    # Pattern: RESUMO DO DIA_M_D_YYYY.xlsx or RESUMO_DO_DIA_M_D_YYYY.xlsx
    match = re.search(r'(\d{1,2})_(\d{1,2})_(\d{4})\.xlsx', filename, re.IGNORECASE)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        year = int(match.group(3))
        try:
            return date(year, month, day)
        except ValueError:
            pass
    return None

def find_estoque_section(ws):
    """
    Find the row where 'Estoque de Combustível' section starts
    Returns the header row number (with Combustível, Inicial, Final, etc.)
    """
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        if cell_value and 'Estoque de Combust' in str(cell_value):
            # The header row is the next row
            return row_num + 1
    return None

def parse_inventory_variance(file_path, company_id):
    """
    Parse the Estoque de Combustível section from RESUMO DO DIA file.
    Returns list of variance records.
    """
    filename = Path(file_path).name
    print(f"Parsing file: {filename}")
    
    # Extract date from filename
    variance_date = extract_date_from_filename(filename)
    if not variance_date:
        print(f"  Warning: Could not extract date from filename: {filename}")
        return []
    
    print(f"  Date: {variance_date}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Find the Estoque de Combustível section
    header_row = find_estoque_section(ws)
    if not header_row:
        print(f"  Warning: Could not find 'Estoque de Combustível' section")
        wb.close()
        return []
    
    print(f"  Found section at row {header_row}")
    
    # Column indices (based on analysis)
    COL_PRODUCT = 0      # Combustível
    COL_INITIAL = 7      # Inicial
    COL_FINAL = 16       # Final
    COL_MEASUREMENT = 26 # Medição
    COL_VARIANCE = 36    # Diferença
    
    records = []
    data_row = header_row + 1  # Data starts after header
    
    # Read until we hit an empty product name or a different section
    while True:
        product_name = ws.cell(row=data_row, column=COL_PRODUCT + 1).value
        
        if not product_name or str(product_name).strip() == '':
            break
        
        product_name = str(product_name).strip().upper()
        
        # Skip if not a fuel product
        if product_name not in PRODUCT_CODE_MAP:
            data_row += 1
            continue
        
        try:
            initial_stock = ws.cell(row=data_row, column=COL_INITIAL + 1).value
            final_stock = ws.cell(row=data_row, column=COL_FINAL + 1).value
            measurement = ws.cell(row=data_row, column=COL_MEASUREMENT + 1).value
            variance = ws.cell(row=data_row, column=COL_VARIANCE + 1).value
            
            # Convert to float, default to 0 if None
            initial_stock = float(initial_stock) if initial_stock is not None else 0
            final_stock = float(final_stock) if final_stock is not None else 0
            measurement = float(measurement) if measurement is not None else 0
            variance = float(variance) if variance is not None else 0
            
            record = {
                'company_id': company_id,
                'variance_date': str(variance_date),
                'product_code': PRODUCT_CODE_MAP[product_name],
                'product_name': product_name,
                'initial_stock': initial_stock,
                'final_stock': final_stock,
                'measurement': measurement,
                'variance': variance,
                'source_file_name': filename
            }
            
            records.append(record)
            print(f"    {product_name}: variance = {variance:+.3f} L")
            
        except Exception as e:
            print(f"  Warning: Error parsing row {data_row}: {e}")
        
        data_row += 1
    
    wb.close()
    print(f"  Parsed {len(records)} variance records")
    return records

def load_to_database(records, supabase):
    """Load parsed records to Supabase database"""
    
    if not records:
        print("No records to load")
        return 0, 0, 0
    
    inserted = 0
    updated = 0
    errors = 0
    
    for record in records:
        try:
            # Check for existing record
            existing = supabase.table('daily_inventory_variance').select('id').eq(
                'company_id', record['company_id']
            ).eq(
                'variance_date', record['variance_date']
            ).eq(
                'product_code', record['product_code']
            ).execute()
            
            if existing.data:
                # Update existing record
                supabase.table('daily_inventory_variance').update({
                    'initial_stock': record['initial_stock'],
                    'final_stock': record['final_stock'],
                    'measurement': record['measurement'],
                    'variance': record['variance'],
                    'source_file_name': record['source_file_name']
                }).eq('id', existing.data[0]['id']).execute()
                updated += 1
            else:
                # Insert new record
                supabase.table('daily_inventory_variance').insert(record).execute()
                inserted += 1
                
        except Exception as e:
            print(f"  Error loading record: {e}")
            errors += 1
    
    print(f"  Results: {inserted} inserted, {updated} updated, {errors} errors")
    return inserted, updated, errors

def main():
    """Main function to run the parser"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Parse inventory variance from RESUMO DO DIA files')
    parser.add_argument('file_path', nargs='?', help='Path to Excel file (optional)')
    parser.add_argument('--company-id', type=int, default=1, help='Company ID (default: 1)')
    parser.add_argument('--all', action='store_true', help='Process all RESUMO DO DIA files in docs folder')
    
    args = parser.parse_args()
    
    supabase = get_supabase_client()
    
    # Process single file or all files
    if args.file_path:
        # Single file mode
        if not Path(args.file_path).exists():
            print(f"Error: File not found: {args.file_path}")
            sys.exit(1)
        
        records = parse_inventory_variance(args.file_path, args.company_id)
        if records:
            load_to_database(records, supabase)
    else:
        # Batch mode - process all RESUMO DO DIA files
        docs_path = Path(__file__).parent.parent.parent / 'docs'
        print(f"Looking in: {docs_path}")
        
        # Find all RESUMO DO DIA files
        pattern = 'RESUMO*DO*DIA*.xlsx'
        files = list(docs_path.glob(pattern)) + list(docs_path.glob(pattern.replace('*', '_')))
        
        # Also try with underscores
        files.extend(docs_path.glob('RESUMO_DO_DIA_*.xlsx'))
        files.extend(docs_path.glob('RESUMO DO DIA_*.xlsx'))
        
        # Remove duplicates
        files = list(set(files))
        files.sort()
        
        if not files:
            print(f"No RESUMO DO DIA files found in {docs_path}")
            sys.exit(1)
        
        print(f"Found {len(files)} file(s) to process\n")
        
        total_inserted = 0
        total_updated = 0
        total_errors = 0
        
        for f in files:
            print(f"\n{'='*50}")
            records = parse_inventory_variance(str(f), args.company_id)
            if records:
                ins, upd, err = load_to_database(records, supabase)
                total_inserted += ins
                total_updated += upd
                total_errors += err
        
        print(f"\n{'='*50}")
        print(f"TOTAL: {total_inserted} inserted, {total_updated} updated, {total_errors} errors")
    
    print("\nDone!")

if __name__ == '__main__':
    main()
