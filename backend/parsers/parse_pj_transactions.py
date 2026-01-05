"""
Parser for Histórico de Consumo (PJ Client Transaction History) Excel files
Extracts transaction-level data for corporate clients (Pessoa Jurídica)
"""

import os
import sys
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Add backend to path
sys.path.append(str(Path(__file__).parent.parent))

# Product name to canonical code mapping
PRODUCT_NAME_TO_CODE = {
    'GASOLINA COMUM': 'GC',
    'GASOLINA COMUM.': 'GC',
    'GASOLINA ADITIVADA': 'GA',
    'GASOLINA ADITIVADA.': 'GA',
    'ETANOL': 'ET',
    'ETANOL.': 'ET',
    'DIESEL S10': 'DS10',
    'DIESEL S10.': 'DS10',
    'DIESEL S-10': 'DS10',
    'DIESEL S500': 'DS500',
    'DIESEL S500.': 'DS500',
    'DIESEL S-500': 'DS500',
    'DIESEL COMUM': 'DS500',
    'DIESEL COMUM.': 'DS500',
}


def parse_client_info(client_string):
    """
    Parse client info from format: "CODE - CNPJ NAME" or "CODE - NAME"
    Examples:
        "008084 - 44.686.944 BRUNNO MENDONCA DE OLIVEIRA PINTO"
        "069957 - 59.671.050 JOSE LEZIO COCO"
        "000001 - CONSUMIDOR FINAL"

    Returns: (client_code, cnpj, client_name)
    """
    if not client_string:
        return None, None, None

    # Split by " - " to separate code from rest
    parts = client_string.split(' - ', 1)
    if len(parts) < 2:
        return None, None, client_string

    client_code = parts[0].strip()
    rest = parts[1].strip()

    # Try to extract CNPJ (format: XX.XXX.XXX or similar patterns)
    # CNPJ pattern: digits with dots, at the start of the string
    cnpj_match = re.match(r'^([\d\.]+)\s+(.+)$', rest)

    if cnpj_match:
        cnpj = cnpj_match.group(1)
        client_name = cnpj_match.group(2).strip()
    else:
        cnpj = None
        client_name = rest

    return client_code, cnpj, client_name


def get_canonical_code(product_name):
    """Map product name to canonical code"""
    if not product_name:
        return None

    name_upper = product_name.upper().strip()
    return PRODUCT_NAME_TO_CODE.get(name_upper)


def parse_date(date_value):
    """Parse date from various formats"""
    if not date_value:
        return None

    # If already a datetime object
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')

    # If string, try to parse
    date_str = str(date_value).strip()

    # Try DD/MM/YYYY format
    try:
        dt = datetime.strptime(date_str, '%d/%m/%Y')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass

    # Try YYYY-MM-DD format
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass

    return None


def parse_historico_consumo(file_path, company_id=2):
    """
    Parse Histórico de Consumo Excel file

    File structure:
    - Row with "Cliente" in col 0: Client info in col 2
    - Row with "Cupom" in col 0: Header row (skip)
    - Row with numeric cupom in col 0: Transaction data
    - Row with "Totais de Data:" in col 0: Date totals (skip)
    - Row with "Total do Cliente:" in col 0: Client totals (skip)

    Column mapping (0-indexed):
    - Col 0: Cupom
    - Col 1: Produto
    - Col 8: Emissão (date)
    - Col 9: Hora
    - Col 16: Quant (volume)
    - Col 18: Val Unit(R$)
    - Col 21: Valor (R$)
    - Col 7: Placa (vehicle plate)
    """

    print(f"[*] Parsing: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    transactions = []
    current_client = None
    current_cnpj = None
    current_client_code = None

    for row in ws.iter_rows(values_only=True):
        cell0 = str(row[0]).strip() if row[0] is not None else ''

        # Check for client row
        if cell0 == 'Cliente':
            client_info = row[2] if len(row) > 2 and row[2] else None
            if client_info:
                current_client_code, current_cnpj, current_client = parse_client_info(str(client_info))
            continue

        # Skip header rows and total rows
        if cell0 in ('Cupom', 'Totais de Data:', 'Total do Cliente:', 'Filial', ''):
            continue

        # Check for data row (numeric cupom)
        if cell0.isdigit() and current_client:
            cupom = cell0
            product_name = str(row[1]).strip() if row[1] else None
            transaction_date = parse_date(row[8])
            volume = float(row[16]) if row[16] else 0
            unit_price = float(row[18]) if row[18] else 0
            total_value = float(row[21]) if row[21] else 0
            vehicle_plate = str(row[7]).strip() if len(row) > 7 and row[7] else None

            # Skip invalid rows
            if not transaction_date or volume == 0:
                continue

            canonical_code = get_canonical_code(product_name)

            transaction = {
                'company_id': company_id,
                'client_code': current_client_code,
                'client_name': current_client,
                'cnpj': current_cnpj,
                'transaction_date': transaction_date,
                'product_name': product_name,
                'canonical_product_code': canonical_code,
                'volume': volume,
                'unit_price': unit_price,
                'total_value': total_value,
                'vehicle_plate': vehicle_plate,
                'cupom': cupom
            }

            transactions.append(transaction)

    wb.close()

    # Summary stats
    pj_count = len([t for t in transactions if 'CONSUMIDOR' not in (t['client_name'] or '').upper()])
    walkin_count = len(transactions) - pj_count
    total_volume = sum(t['volume'] for t in transactions)

    print(f"    Total transactions: {len(transactions)}")
    print(f"    PJ transactions: {pj_count}")
    print(f"    Walk-in transactions: {walkin_count}")
    print(f"    Total volume: {total_volume:,.2f} L")

    return transactions


def load_transactions_to_database(transactions, include_walkin=True):
    """
    Load transactions to Supabase database

    Args:
        transactions: List of transaction dicts
        include_walkin: If True, include CONSUMIDOR FINAL transactions
    """
    from dotenv import load_dotenv
    from supabase import create_client

    # Load environment
    env_path = Path(__file__).parent.parent.parent / 'config' / '.env.local'
    load_dotenv(env_path)

    supabase_url = os.getenv('SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_SERVICE_KEY')

    if not supabase_url or not supabase_key:
        raise ValueError("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY in environment")

    supabase = create_client(supabase_url, supabase_key)

    inserted = 0
    skipped = 0
    errors = []

    for trans in transactions:
        # Filter out walk-in if requested
        if not include_walkin and 'CONSUMIDOR' in (trans['client_name'] or '').upper():
            continue

        try:
            # Check for existing record
            existing = supabase.table('pj_client_transactions').select('id').eq(
                'company_id', trans['company_id']
            ).eq(
                'cupom', trans['cupom']
            ).eq(
                'transaction_date', trans['transaction_date']
            ).eq(
                'client_code', trans['client_code']
            ).execute()

            if existing.data:
                skipped += 1
                continue

            # Insert new record
            supabase.table('pj_client_transactions').insert(trans).execute()
            inserted += 1

        except Exception as e:
            errors.append({
                'cupom': trans['cupom'],
                'error': str(e)
            })

    return {
        'inserted': inserted,
        'skipped': skipped,
        'errors': errors
    }


def main():
    """Main entry point for testing with a single file"""
    import argparse

    parser = argparse.ArgumentParser(description='Parse Histórico de Consumo files')
    parser.add_argument('--file', type=str, help='Path to Excel file')
    parser.add_argument('--company-id', type=int, default=2, help='Company ID (default: 2)')
    parser.add_argument('--load', action='store_true', help='Load to database')
    parser.add_argument('--include-walkin', action='store_true', help='Include walk-in transactions')

    args = parser.parse_args()

    # Default to December file if not specified
    if not args.file:
        docs_path = Path(__file__).parent.parent.parent / 'docs'
        args.file = str(docs_path / 'Histórico de Consumo.xlsx')

    if not os.path.exists(args.file):
        print(f"[X] File not found: {args.file}")
        return

    # Parse file
    transactions = parse_historico_consumo(args.file, args.company_id)

    if args.load:
        print("\n[*] Loading to database...")
        result = load_transactions_to_database(transactions, args.include_walkin)
        print(f"    Inserted: {result['inserted']}")
        print(f"    Skipped: {result['skipped']}")
        if result['errors']:
            print(f"    Errors: {len(result['errors'])}")
            for err in result['errors'][:3]:
                print(f"      - Cupom {err['cupom']}: {err['error']}")

    print("\n[OK] Done!")


if __name__ == '__main__':
    main()
